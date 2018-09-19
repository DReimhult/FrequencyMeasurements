# -*- coding: utf-8 -*-
"""
Created on Thu Mar 29 15:26:26 2018

@author: Daniel Reimhult

@version: 0.1
"""

import numpy as np
from openpyxl import load_workbook
    
class Signal:
    def __init__(self, time, values, sample_rate, nom_freq = 0):
        self.time = time
        self.values = values
        self.sample_rate = sample_rate
        self.nom_freq = nom_freq

        if nom_freq > 0:
            self.samples_per_period = sample_rate // nom_freq
      
    def __add__(self, other):
        retsig = self
        retsig.values = [a+b for a,b in zip(self.values,other.values)]
        return retsig
        
    def plot(self, ax, **kwargs):
        ax.plot(self.time, self.values, **kwargs)
                
def generate_test_signal( fnom = 50, num_cycles = 100, sample_rate = 20000, \
                          amplitude = 1, variations = [], disturbances = [] ):
    """ Returns test signal for frequency method testing. """
    
    spp = sample_rate//fnom
    ss = np.arange(0,spp*num_cycles)
    N = len(ss)
    time = [s / sample_rate for s in ss]

    signal = [amplitude * np.sin(2*np.pi*fnom*tp) for tp in time]

    for v in variations:
        if v['name'] == 'constant':
            pass
        elif v['name'] == 'phase jump':
            startidx = int(spp*v['params'][0])
            endidx = startidx + int(spp*v['params'][1])
            signal[startidx:endidx] = [amplitude * np.sin(2*np.pi*fnom*tp + \
                             v['params'][2]) for tp in time[startidx:endidx]]
        elif v['name'] == 'slow mod':
            signal = [amplitude * np.sin(2*np.pi*fnom*tp + \
                                    0.2*np.sin(2*np.pi*0.2*tp)) for tp in time]
        elif v['name'] == 'phase jump and slow mod':
            signal = [amplitude * np.sin(2*np.pi*fnom*tp + \
                                    0.5*np.sin(2*np.pi*0.2*tp)) for tp in time]
            startidx = int(spp*v['params'][0])
            endidx = startidx + int(spp*v['params'][1])
            signal[startidx:endidx] = [amplitude * np.sin(2*np.pi*fnom*tp + \
                                    0.5*np.sin(2*np.pi*0.2*tp) + \
                                    v['params'][2]) for tp in time[startidx:endidx]]
    
    for d in disturbances:
        if d['name'] == 'white noise':
            r = np.random.normal(0.0, d['params'][0], size=N)
            signal = [a+b for a,b in zip(signal,r)]
        elif d['name'] == 'superposed frequency':
            r = [d['params'][0] * np.sin(2*np.pi*d['params'][1]*tp) \
                 for tp in time]
            signal = [a+b for a,b in zip(signal,r)]
        elif d['name'] == 'harmonic':
            r = [d['params'][0] * np.sin(2*np.pi*d['params'][1]*fnom*tp + \
                 d['params'][2]) for tp in time]
            signal = [a+b for a,b in zip(signal,r)]
        else:
            raise ValueError('Disturbance \"{dn}\" not supported.' \
                             .format(dn=d['name']))
            
    return Signal(time,signal,sample_rate,fnom)

def read_ref_signal_from_xls(filename):
    wb = load_workbook(filename, data_only=True)
    
    ws = wb['Sheet1']
    
    signal = ws['K2':'L67044']
    spp = ws['G2'].value
    frequency = ws['G17'].value
    sample_rate = spp * frequency
    
    t = [s[0].value for s in signal]
    data = [s[1].value for s in signal]
    
    return Signal(t,data,sample_rate,frequency)
